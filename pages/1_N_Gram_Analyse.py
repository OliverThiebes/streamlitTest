import io
import re
import csv
from typing import Dict, List, Tuple
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

st.set_page_config(page_title="N-Gram Analyse", layout="centered")

# Kleinbuchstaben, Token inkl. Umlaute und +/-
TOKEN_RE = re.compile(r"[a-z0-9\+\-äüöß]+", re.IGNORECASE)

# Anzahl Beispiel-Suchbegriffe je Gram
TOP_TERMS_LIMIT = 5

# Alle n-grams von 1 bis 5 Wörtern berücksichtigen
NGRAM_RANGE = range(1, 6)

# EUR-Umrechnungskurse (1 Einheit Fremdwährung -> EUR), Stand 21.01.2026
# Quelle: ECB/TradingEconomics Tagesreferenzkurse.
FX_RATES_TO_EUR: Dict[str, float] = {
    "EUR": 1.0,
    "USD": 0.85186,
    "GBP": 1.14364,
    "CHF": 1.07898,
    "DKK": 0.13386,
    "NOK": 0.08594,
    "SEK": 0.09380,
    "PLN": 0.23655,
    "CZK": 0.04106,
}


def sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t"])
        return dialect.delimiter
    except Exception:
        return ";"


def read_google_ads_csv_bytes(raw: bytes) -> pd.DataFrame:
    """
    Google Ads CSV hat oft 2 Info-Zeilen am Anfang.
    Überschriften stehen laut dir in Zeile 3.
    """
    # Google Ads Exporte sind oft UTF-16, manchmal UTF-8 mit BOM
    for enc in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("latin-1")

    sample = "\n".join(text.splitlines()[:20])
    delimiter = sniff_delimiter(sample)

    from io import StringIO

    df = pd.read_csv(
        StringIO(text),
        sep=delimiter,
        skiprows=2,  # Zeile 1-2 uninteressant
        dtype=str,
        engine="python",
    )

    # Spalten bereinigen
    df.columns = [c.strip() for c in df.columns]
    return df


def _parse_number_de(t: str) -> float:
    """Hilfsfunktion: deutsches Zahlenformat ohne Währung in float."""
    t = t.replace(".", "")
    t = t.replace("\u00a0", " ")
    t = t.replace(" ", "")
    t = t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def parse_de_number(s: str) -> float:
    """Parsing für nicht-monetäre Spalten (Impressions, Klicks, Conversions)."""
    if s is None:
        return 0.0
    t = str(s).strip()
    if t == "" or t.lower() in ("--", "null", "nan"):
        return 0.0
    t = re.sub(r"[^\d.,\- ]", "", t)
    return _parse_number_de(t)


def parse_currency_to_eur(s: str) -> float:
    """
    Ermittelt Währung per Symbol/Code und wandelt in EUR um.
    Unterstützt: EUR, USD ($), GBP (£), CHF (CHF/Fr), DKK, NOK, SEK, PLN, CZK.
    """
    if s is None:
        return 0.0
    raw = str(s)
    if raw.strip() == "" or raw.lower() in ("--", "null", "nan"):
        return 0.0

    up = raw.upper()
    if "$" in raw or "USD" in up:
        cur = "USD"
    elif "£" in raw or "GBP" in up:
        cur = "GBP"
    elif "CHF" in up or "FR" in up:
        cur = "CHF"
    elif "DKK" in up:
        cur = "DKK"
    elif "NOK" in up:
        cur = "NOK"
    elif "SEK" in up:
        cur = "SEK"
    elif "PLN" in up:
        cur = "PLN"
    elif "CZK" in up:
        cur = "CZK"
    else:
        cur = "EUR"

    number_part = re.sub(r"[^\d.,\- ]", "", raw)
    val = _parse_number_de(number_part)
    rate = FX_RATES_TO_EUR.get(cur, 1.0)
    return val * rate


def tokenize_search_term(search_term: str) -> List[str]:
    if search_term is None:
        return []
    lower = str(search_term).lower()
    return TOKEN_RE.findall(lower)


def format_float_de(x: float, digits: int = 1) -> str:
    s = f"{x:.{digits}f}"
    return s.replace(".", ",")


def build_ngrams(tokens: List[str], n: int) -> List[str]:
    if n <= 0 or len(tokens) < n:
        return []
    if n == 1:
        return tokens
    return [" ".join(tokens[i : i + n]) for i in range(len(tokens) - n + 1)]


def transform_to_ngram_table(
    df: pd.DataFrame,
    progress_cb=None,
    label: str = "",
) -> Tuple[pd.DataFrame, Dict[str, Dict[str, float]]]:
    required = {
        "Suchbegriff",
        "Impressionen",
        "Klicks",
        "Kosten",
        "Conversions",
        "Conv.-Wert",
    }
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            "Fehlende Spalten in der CSV: "
            + ", ".join(missing)
            + ". Prüfe, ob die Überschriften exakt wie angegeben sind."
        )

    # Numerische Spalten parse
    impressions = df["Impressionen"].map(parse_de_number)
    clicks = df["Klicks"].map(parse_de_number)
    cost = df["Kosten"].map(parse_currency_to_eur)
    conversions = df["Conversions"].map(parse_de_number)
    conv_value = df["Conv.-Wert"].map(parse_currency_to_eur)

    agg: Dict[str, Dict[str, float]] = {}
    examples: Dict[str, Dict[str, float]] = {}
    occurrences: Dict[str, int] = {}
    term_stats: Dict[str, Dict[str, float]] = {}

    total_rows = len(df)
    for i in range(total_rows):
        if progress_cb and total_rows > 0 and (i % 50 == 0 or i == total_rows - 1):
            progress_cb(i + 1, total_rows, label)

        st_val = df.at[i, "Suchbegriff"]
        # Google Ads Gesamt-Zeilen (Summen) überspringen
        if st_val is not None and str(st_val).strip().lower().startswith("gesamt"):
            continue
        tokens = tokenize_search_term(st_val)
        grams: List[str] = []
        for n in NGRAM_RANGE:
            grams.extend(build_ngrams(tokens, n))

        unique_grams = set(grams)
        for gram in unique_grams:
            if gram not in agg:
                agg[gram] = {
                    "impressions": 0.0,
                    "clicks": 0.0,
                    "cost": 0.0,
                    "conversions": 0.0,
                    "conversionValue": 0.0,
                }
                examples[gram] = {}
                occurrences[gram] = 0

            agg[gram]["impressions"] += float(impressions.iat[i])
            agg[gram]["clicks"] += float(clicks.iat[i])
            agg[gram]["cost"] += float(cost.iat[i])
            agg[gram]["conversions"] += float(conversions.iat[i])
            agg[gram]["conversionValue"] += float(conv_value.iat[i])
            occurrences[gram] += 1

            if st_val:
                examples[gram][str(st_val)] = examples[gram].get(str(st_val), 0.0) + float(
                    impressions.iat[i]
                )

        # Term-Stats (exakter Suchbegriff)
        if st_val:
            tkey = str(st_val)
            if tkey not in term_stats:
                term_stats[tkey] = {"cost": 0.0, "conv_value": 0.0}
            term_stats[tkey]["cost"] += float(cost.iat[i])
            term_stats[tkey]["conv_value"] += float(conv_value.iat[i])

    rows: List[Dict[str, object]] = []
    for gram, m in agg.items():
        imps = m["impressions"]
        clks = m["clicks"]
        cst = m["cost"]
        convs = m["conversions"]
        convv = m["conversionValue"]

        ctr = (clks / imps) if imps > 0 else 0.0
        roas = (convv / cst) if cst > 0 else 0.0
        kur = (cst / convv) if convv > 0 else None

        top_examples: List[Dict[str, object]] = []
        if imps > 0:
            sorted_terms = sorted(
                examples.get(gram, {}).items(),
                key=lambda x: x[1],
                reverse=True,
            )[:TOP_TERMS_LIMIT]
            for term, imp_term in sorted_terms:
                share = (imp_term / imps) * 100
                top_examples.append({"term": term, "share": share})
        else:
            for term in list(examples.get(gram, {}).keys())[:TOP_TERMS_LIMIT]:
                top_examples.append({"term": term, "share": 0.0})

        rows.append(
            {
                "gram": gram,
                "gram_count": len(gram.split()),
                "vorkommen": occurrences.get(gram, 0),
                "_examples": top_examples,
                "impressions": imps,
                "clicks": clks,
                "ctr": ctr,
                "cost": cst,
                "conversions": convs,
                "conversionValue": convv,
                "ROAS": roas,
                "KUR": kur,
            }
        )

    out = pd.DataFrame(rows)

    # Sortierung: nach cost absteigend
    if not out.empty:
        out = out.sort_values(by="cost", ascending=False)

    # Spaltenreihenfolge (top_search_terms wird später befüllt)
    cols = [
        "gram",
        "gram_count",
        "vorkommen",
        "top_search_terms",
        "_examples",
        "impressions",
        "clicks",
        "ctr",
        "cost",
        "conversions",
        "conversionValue",
        "ROAS",
        "KUR",
    ]
    out = out.reindex(columns=cols)

    return out, term_stats


def apply_number_formats_and_rules(ws, row_count: int) -> None:
    if row_count == 0:
        return

    # Nummernformate (Zeilen 2..row_count+1)
    formats = {
        5: "#,##0",  # impressions
        6: "#,##0",  # clicks
        7: "0.00%",  # ctr
        8: '#,##0.00 "€"',  # cost
        9: "0.00",  # conversions
        10: '#,##0.00 "€"',  # conversionValue
        11: "0.00%",  # ROAS
        12: "0.00%",  # KUR
    }
    for row in ws.iter_rows(min_row=2, max_row=row_count + 1, min_col=5, max_col=12):
        for cell in row:
            fmt = formats.get(cell.col_idx)
            if fmt:
                cell.number_format = fmt

    # Bedingte Formatierung: ganze Zeile rot bei Kosten >100 EUR und KUR >45%
    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    start_row = 2
    end_row = row_count + 1
    data_range = f"A{start_row}:L{end_row}"
    rule = FormulaRule(formula=["AND($H2>100,$L2>0.45)"], fill=fill)
    ws.conditional_formatting.add(data_range, rule)


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/*?\\[\\]]", "_", name)
    return cleaned[:31] or "Sheet"


st.title("N-Gram Analyse")
st.write(
    "High-Level: Du lädst eine oder mehrere Google-Ads-Suchbegriffe-CSV(s) hoch. "
    "Die App tokenisiert die Suchbegriffe, erzeugt N-Grams (1–5 Wörter), "
    "aggregiert Leistungswerte (Impressions, Klicks, Kosten, Conversions, Wert) "
    "und exportiert eine Excel-Datei mit einem Sheet pro Upload."
)
st.markdown(
    "Erwartete Spalten (exakt): "
    "`Suchbegriff`, `Impressionen`, `Klicks`, `Kosten`, `Conversions`, `Conv.-Wert`."
)

uploaded_files = st.file_uploader(
    "Google Ads CSV(s) hochladen",
    type=["csv"],
    accept_multiple_files=True,
)

if uploaded_files:
    run = st.button("Analyse starten")
    if run:
        progress = st.progress(0)
        status = st.empty()

        date_prefix = datetime.now().strftime("%Y%m%d")
        out_name = f"{date_prefix}_JHPS_N-Gram.xlsx"

        out_buf = io.BytesIO()
        with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
            for idx, up in enumerate(uploaded_files, start=1):
                try:
                    status.write(f"Verarbeite Datei {idx}/{len(uploaded_files)}: {up.name}")

                    def progress_cb(done: int, total: int, label: str) -> None:
                        base = (idx - 1) / len(uploaded_files)
                        part = (done / total) / len(uploaded_files) if total else 0
                        progress.progress(min(base + part, 1.0))

                    df = read_google_ads_csv_bytes(up.getvalue())
                    out_df, term_stats = transform_to_ngram_table(
                        df,
                        progress_cb=progress_cb,
                        label=up.name,
                    )
                    export_df = out_df[out_df["cost"] > 0].reset_index(drop=True)

                    # Rote Grammata (Kosten>100 & KUR>45%)
                    red_grams = set(
                        export_df.loc[(export_df["cost"] > 100) & (export_df["KUR"] > 0.45), "gram"]
                    )

                    def canonical(term: str) -> str:
                        return " ".join(tokenize_search_term(term))

                    def build_examples_cell(row) -> str:
                        parts: List[str] = []
                        for ex in row["_examples"]:
                            term = ex["term"]
                            share = ex["share"]
                            can = canonical(term)
                            is_red = (term in red_grams) or (can in red_grams)
                            mark = " ❌" if is_red else ""
                            extra = ""
                            if is_red:
                                info_term = term_stats.get(term)
                                if info_term:
                                    cost_val = info_term["cost"]
                                    conv_val = info_term["conv_value"]
                                    kur_val = (cost_val / conv_val) if conv_val > 0 else None
                                    cost_txt = f"{format_float_de(cost_val, 2)} €"
                                    kur_txt = (
                                        f"{format_float_de(kur_val * 100, 1)} %"
                                        if kur_val is not None
                                        else ""
                                    )
                                    extra = f" [Kosten: {cost_txt}; KUR: {kur_txt}]"
                            parts.append(
                                f"{term} ({format_float_de(share, 1)} %){mark}{extra}"
                            )
                        return ", ".join(parts)

                    export_df["top_search_terms"] = export_df.apply(build_examples_cell, axis=1)
                    export_df = export_df.drop(columns=["_examples"])

                    sheet = safe_sheet_name(up.name.rsplit(".", 1)[0])
                    export_df.to_excel(writer, index=False, sheet_name=sheet)

                    ws = writer.sheets[sheet]
                    apply_number_formats_and_rules(ws, len(export_df))
                except Exception as e:
                    msg = str(e)
                    if "Fehlende Spalten" in msg:
                        st.error(
                            f"Datei '{up.name}' hat nicht die erwarteten Spalten. "
                            "Bitte prüfe die Überschriften."
                        )
                    else:
                        st.error(
                            f"Datei '{up.name}' konnte nicht verarbeitet werden. "
                            f"Details: {msg}"
                        )

        out_buf.seek(0)
        progress.progress(1.0)
        status.write("Fertig.")

        st.download_button(
            "Excel herunterladen",
            data=out_buf.getvalue(),
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.caption("Bitte eine oder mehrere CSV-Dateien hochladen.")
